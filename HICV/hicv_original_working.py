import os
import re
import csv
import time
from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PwTimeout
from dotenv import load_dotenv

# ----------------- Config -----------------
load_dotenv()
USERNAME = os.getenv("HICV_USERNAME")
PASSWORD = os.getenv("HICV_PASSWORD")

if not USERNAME or not PASSWORD:
    print("❌ ERROR: Missing credentials in .env file.")
    exit(1)

# 🔧 Test controls
PROMPT_FOR_INPUT = False          # set True later when you want prompts again
TEST_MONTH = "January"
TEST_YEAR = "2026"
CHECKIN_DAY = 1                   # for our current test: “first of the month”

ADULTS = 2
CHILDREN = 0
UNIT_SIZES = ["Studio", "1 Bedroom"]
NIGHTS = 7                        # we’ll try to select an end date after start

CSV_PATH = Path("member_availability.csv")
XLSX_PATH = Path("member_availability.xlsx")
HEADLESS = False
SLOW_MO = 200

# ----------------- Input helpers -----------------
MONTHS = {
    "january":"January","february":"February","march":"March","april":"April","may":"May","june":"June",
    "july":"July","august":"August","september":"September","october":"October","november":"November","december":"December"
}
MONTH_SHORT = {
    "January":"Jan","February":"Feb","March":"Mar","April":"Apr","May":"May","June":"Jun",
    "July":"Jul","August":"Aug","September":"Sep","October":"Oct","November":"Nov","December":"Dec"
}

def prompt_month_year():
    if not PROMPT_FOR_INPUT:
        month = TEST_MONTH
        year = TEST_YEAR
        checkin_full_aria = f"{month} {CHECKIN_DAY}, {year}"
        print(f"🔧 Start at: {month} {year} — target day {CHECKIN_DAY}")
        return month, year, checkin_full_aria

    m_raw = input(f"Enter month (e.g., December) [{TEST_MONTH}]: ").strip()
    y_raw = input(f"Enter year (e.g., 2026) [{TEST_YEAR}]: ").strip()

    month_key = (m_raw or TEST_MONTH).strip().lower()
    month = MONTHS.get(month_key)
    if not month:
        raise ValueError(f"Unsupported month: {m_raw!r}. Use full month name (e.g., December).")

    year = (y_raw or TEST_YEAR).strip()
    if not year.isdigit() or not (1900 <= int(year) <= 2100):
        raise ValueError(f"Invalid year: {y_raw!r}. Use a 4-digit year like 2026.")

    checkin_full_aria = f"{month} {CHECKIN_DAY}, {year}"
    return month, year, checkin_full_aria

# ----------------- Common helpers -----------------
def write_debug(page, stem: str):
    try:
        page.screenshot(path=f"{stem}.png", full_page=True)
    except Exception:
        pass
    try:
        Path(f"{stem}.html").write_text(page.content(), encoding="utf-8")
    except Exception:
        pass

def accept_cookies_if_present(scope):
    try:
        scope.get_by_role("button", name=re.compile(r"^I Agree$", re.I)).click(timeout=3000)
        time.sleep(0.3)
        print("🍪 Cookie banner accepted.")
    except PwTimeout:
        pass

def wait_for_booking_form(page, timeout=25000):
    try:
        page.locator("#resorts-dropdown").wait_for(state="visible", timeout=timeout)
    except Exception:
        page.locator("text=Select Check In Date").wait_for(timeout=timeout)

def wait_for_results(page, timeout=30000):
    anchors = [
        page.locator("[data-uitest='availability-results']").first,
        page.locator("h1:has-text('Book a Villa')"),
        page.locator("text=Modify Search"),
        page.locator("h2"),
    ]
    last_err = None
    for a in anchors:
        try:
            a.wait_for(state="visible", timeout=timeout)
            return
        except Exception as e:
            last_err = e
    write_debug(page, "results_wait_failed")
    raise last_err

def open_resorts_dropdown(page):
    btn = page.locator(
        "#resorts-dropdown, button[data-uitest='multi-select-button'], [aria-label='Open multi select.']"
    ).first
    btn.wait_for(state="visible", timeout=15000)
    btn.scroll_into_view_if_needed()

    panel_sel = "ul[data-uitest='ul-resorts'][role='listbox'], ul[role='listbox']"
    panel = page.locator(panel_sel).first

    for attempt in range(6):
        try:
            if attempt == 0: btn.click(timeout=1500)
            elif attempt == 1: btn.click(force=True, timeout=1500)
            elif attempt == 2: btn.dblclick(timeout=1200)
            elif attempt == 3: btn.focus(); page.keyboard.press("Enter")
            elif attempt == 4: btn.focus(); page.keyboard.press("Space")
            else:
                eh = btn.element_handle()
                if eh:
                    page.evaluate("(el)=>el.dispatchEvent(new MouseEvent('click',{bubbles:true}))", eh)
            panel.wait_for(state="visible", timeout=1500)
            return panel
        except Exception:
            time.sleep(0.15)

    write_debug(page, "open_dropdown_failed")
    raise TimeoutError("Location listbox did not appear after clicking the dropdown.")

# ---------- Florida selection helpers ----------
def _collect_florida_ids_from_dom(panel):
    return panel.evaluate(
        """
        (el) => {
          const items = Array.from(el.querySelectorAll("li[role='option']"));
          const isGroup = li => (li.className || "").includes("group-label");
          const startIdx = items.findIndex(li => li.id === "list-item-Florida-id");
          if (startIdx < 0) return { ids: [], has_next_group: false };

          const ids = [];
          let has_next_group = false;
          for (let i = startIdx + 1; i < items.length; i++) {
            const li = items[i];
            if (isGroup(li)) { has_next_group = true; break; }
            const input = li.querySelector("input[type='checkbox']");
            if (input?.id) ids.push(input.id);
          }
          return { ids, has_next_group };
        }
        """
    )

def select_all_florida_locations(page):
    panel = open_resorts_dropdown(page)

    # Scroll until Florida header visible
    found_florida = False
    for _ in range(60):
        if panel.locator("#list-item-Florida-id").count() > 0:
            found_florida = True
            break
        try:
            panel.evaluate("el => el.scrollTop = Math.min(el.scrollTop + el.clientHeight, el.scrollHeight)")
        except Exception:
            pass
        time.sleep(0.08)
    if not found_florida:
        write_debug(page, "location_panel_no_florida")
        raise Exception("Couldn't find the Florida group header in the list.")

    # Ensure all Florida options rendered
    all_ids = set()
    has_next = False
    for _ in range(80):
        snap = _collect_florida_ids_from_dom(panel)
        for fid in snap["ids"]:
            all_ids.add(fid)
        has_next = snap["has_next_group"]
        if has_next and len(all_ids) > 0:
            break
        try:
            panel.evaluate("el => el.scrollTop = Math.min(el.scrollTop + Math.floor(el.clientHeight/2), el.scrollHeight)")
        except Exception:
            pass
        time.sleep(0.08)

    for_ids = list(all_ids)
    if not for_ids:
        write_debug(page, "location_panel_after")
        raise Exception("Florida group found, but no Florida options were rendered.")

    newly_selected = 0
    for fid in for_ids:
        li = panel.locator(f"//input[@id='{fid}']/ancestor::li[1]").first
        selected = (li.get_attribute("aria-selected") or "").lower() == "true"
        if selected:
            continue

        label = panel.locator(f"label[for='{fid}']").first
        try:
            page.evaluate("(el)=>el.scrollIntoView({block:'center'})", label.element_handle())
        except Exception:
            pass

        try:
            label.click(timeout=800)
        except Exception:
            try:
                label.click(force=True, timeout=800)
            except Exception:
                li.click(force=True, timeout=800)

        for _ in range(12):
            selected = (li.get_attribute("aria-selected") or "").lower() == "true"
            if selected:
                newly_selected += 1
                break
            time.sleep(0.05)

        if not selected:
            page.evaluate(
                """
                (id) => {
                  const input = document.getElementById(id);
                  if (!input) return false;
                  const li = input.closest("li[role='option']");
                  const desc = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'checked');
                  if (desc && desc.set) desc.set.call(input, true);
                  else input.checked = true;
                  input.dispatchEvent(new Event('input', { bubbles: true }));
                  input.dispatchEvent(new Event('change', { bubbles: true }));
                  if (li) li.setAttribute('aria-selected', 'true');
                  return true;
                }
                """,
                fid,
            )
            time.sleep(0.1)
            selected = (li.get_attribute("aria-selected") or "").lower() == "true"
            if selected:
                newly_selected += 1

    confirmed = 0
    for fid in for_ids:
        li = panel.locator(f"//input[@id='{fid}']/ancestor::li[1]").first
        if (li.get_attribute("aria-selected") or "").lower() == "true":
            confirmed += 1

    print(f"📍 Florida options under header: {len(for_ids)}")
    print(f"✅ Newly selected this run: {newly_selected}")
    print(f"✅ Confirmed selected now: {confirmed}")

    try:
        panel.screenshot(path="location_panel_after.png")
        print("📸 Saved location_panel_after.png")
    except Exception:
        pass

    page.keyboard.press("Escape")
    time.sleep(0.2)
    btn = page.locator("#resorts-dropdown").first
    for _ in range(20):
        try:
            txt = (btn.locator(".FormFieldButton_text__jmMUQ").inner_text() or "").strip()
        except Exception:
            txt = ""
        if txt and ("Selected" in txt or txt.isdigit()):
            break
        time.sleep(0.1)
    try:
        btn.screenshot(path="location_button_after.png")
        print("📸 Saved location_button_after.png")
    except Exception:
        pass

# ---------- Counters & Unit Size ----------
def set_counter(page, kind: str, target: int):
    if kind == "adults":
        span = page.locator("[data-uitest='number-of-adults-number-of-guests']").first
        minus_btn = page.locator("[data-uitest='number-of-adults-left-icon-button']").first
        plus_btn  = page.locator("[data-uitest='number-of-adults-right-icon-button']").first
    else:
        span = page.locator("[data-uitest='number-of-children-number-of-guests']").first
        minus_btn = page.locator("[data-uitest='number-of-children-left-icon-button']").first
        plus_btn  = page.locator("[data-uitest='number-of-children-right-icon-button']").first

    cur = int((span.inner_text() or "0").strip())
    tries = 0
    while cur != target and tries < 30:
        if cur < target:
            plus_btn.click()
        else:
            try:
                minus_btn.click()
            except Exception:
                pass
        time.sleep(0.1)
        try:
            cur = int((span.inner_text() or "0").strip())
        except Exception:
            pass
        tries += 1

def set_unit_sizes(page, desired_labels):
    page.locator("#unit-size-dropdown").click()
    panel = page.locator("ul[data-uitest='ul-unit-sizes']").first
    panel.wait_for(state="visible", timeout=5000)

    options = [
        ("Studio",      "option-ST-id"),
        ("1 Bedroom",   "option-1BD-id"),
        ("2 Bedroom",   "option-2BD-id"),
        ("3+ Bedroom",  "option-3BDPlus-id"),
    ]

    def ensure_state(name, input_id, should_select: bool):
        row_label = panel.locator(f"li[role='option'] >> label[for='{input_id}']").first
        if row_label.count() == 0:
            return
        li = panel.locator(f"//input[@id='{input_id}']/ancestor::li[1]").first
        selected = (li.get_attribute("aria-selected") or "").lower() == "true"
        if selected == should_select:
            return

        try:
            row_label.click(timeout=800)
        except Exception:
            try:
                row_label.click(force=True, timeout=800)
            except Exception:
                li.click(force=True, timeout=800)

        for _ in range(12):
            selected = (li.get_attribute("aria-selected") or "").lower() == "true"
            if selected == should_select:
                return
            time.sleep(0.05)

        page.evaluate(
            """
            ([id, should]) => {
              const input = document.getElementById(id);
              if (!input) return false;
              const li = input.closest("li[role='option']");
              const desc = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'checked');
              if (desc && desc.set) desc.set.call(input, should);
              else input.checked = should;
              input.dispatchEvent(new Event('input', { bubbles: true }));
              input.dispatchEvent(new Event('change', { bubbles: true }));
              if (li) li.setAttribute('aria-selected', should ? 'true' : 'false');
              return true;
            }
            """,
            [input_id, should_select],
        )
        time.sleep(0.1)

    for name, input_id in options:
        ensure_state(name, input_id, name in desired_labels)

    final_selected = []
    for name, input_id in options:
        li = panel.locator(f"//input[@id='{input_id}']/ancestor::li[1]").first
        sel = (li.get_attribute("aria-selected") or "").lower() == "true"
        if sel:
            final_selected.append(name)
    print(f"✅ Unit sizes selected: {final_selected}")

    page.keyboard.press("Escape")
    time.sleep(0.15)

# ---------- Month/Year dropdowns ----------
def pick_month(page, month_full):
    short = MONTH_SHORT[month_full]
    page.locator("#month-picker").click()
    month_btn = page.locator(f"p[role='button']:has-text('{short}')").first
    if month_btn.count() == 0:
        month_btn = page.locator(f"text={short}").first
    month_btn.click()

def pick_year(page, year):
    page.locator("#year-picker").click()
    # Try multiple strategies for year selection
    for sel in (
        f"label[for='year-selection-{year}']",
        f"#year-selection-{year}",
        f"p[role='button']:has-text('{year}')",
        f"//p[normalize-space()='{year}']",
    ):
        opt = page.locator(sel).first
        if opt.count():
            try:
                opt.click(timeout=1200); return
            except Exception:
                opt.click(force=True, timeout=1200); return
    try:
        page.get_by_role("option", name=str(year)).click(timeout=1200); return
    except Exception:
        pass
    raise RuntimeError(f"Could not select year {year}")

# ---------- Calendar helpers ----------
def _open_calendar(page):
    overlay = page.locator("#range-picker, .rdp").first
    if overlay.count() and overlay.is_visible():
        return
    openers = [
        "#date-picker",
        "button[data-uitest='date-picker']",
        "button:has-text('Select Check-In Date')",
        "button[aria-label*='Check-In Date']",
    ]
    for sel in openers:
        btn = page.locator(sel).first
        if btn.count() == 0:
            continue
        try:
            btn.scroll_into_view_if_needed()
        except Exception:
            pass
        for _ in range(2):
            try:
                btn.click(timeout=1200); break
            except Exception:
                btn.click(force=True, timeout=1200)
        try:
            overlay.wait_for(state="visible", timeout=2000)
            return
        except Exception:
            pass
    raise RuntimeError("Could not open the check-in calendar overlay.")

def _wait_calendar_ready(page, timeout_ms=9000):
    overlay = page.locator("#range-picker, .rdp").first
    overlay.wait_for(state="visible", timeout=timeout_ms)
    start = time.time()
    while (time.time() - start) * 1000 < timeout_ms:
        caps = page.locator(".Calendar_month-text__eEKDO")
        enabled = page.locator(
            "div.rdp-month button[class*='rdp-day']:not([disabled]), "
            "div.rdp-month button[data-uitest*='calendar-day']:not([disabled])"
        )
        if caps.count() > 0 and enabled.count() > 0:
            return True
        time.sleep(0.15)
    return False

def _find_day_btn_in_block(mb, day_num: int):
    num = mb.locator(f"span[aria-hidden='true']:text-is('{day_num}')")
    if num.count() == 0:
        return None
    btn = num.locator("xpath=ancestor::button[1]")
    if btn.count() == 0:
        return None
    if btn.get_attribute("disabled") is not None:
        return None
    return btn

def _first_enabled_after(mb, min_day: int):
    for d in range(min_day + 1, 32):
        btn = _find_day_btn_in_block(mb, d)
        if btn:
            return btn
    return None

def _try_click_range(page, start_mb, start_day: int):
    start_btn = _find_day_btn_in_block(start_mb, start_day)
    if not start_btn:
        return False
    try: start_btn.scroll_into_view_if_needed()
    except Exception: pass
    try: start_btn.click(timeout=1500)
    except Exception: start_btn.click(force=True, timeout=1500)

    target_end = start_day + max(1, NIGHTS - 1)
    end_btn = _find_day_btn_in_block(start_mb, target_end)
    if not end_btn:
        end_btn = _first_enabled_after(start_mb, start_day)

    if not end_btn:
        blocks = page.locator("div.rdp-month")
        other = blocks.nth(1) if start_mb == blocks.nth(0) else blocks.nth(0)
        if other.count():
            exact = _find_day_btn_in_block(other, (target_end - 31) if target_end > 31 else target_end)
            end_btn = exact or other.locator("button[class*='rdp-day']:not([disabled])").first

    if not end_btn or end_btn.count() == 0:
        return False

    try: end_btn.scroll_into_view_if_needed()
    except Exception: pass
    try: end_btn.click(timeout=1500)
    except Exception: end_btn.click(force=True, timeout=1500)

    return True

def pick_checkin_date(page, month_full, year, day):
    _open_calendar(page)
    if not _wait_calendar_ready(page, timeout_ms=10000):
        try:
            Path("calendar_dump_before_ready.html").write_text(
                page.locator("#range-picker, .rdp").first.evaluate("el=>el.outerHTML"),
                encoding="utf-8"
            )
        except Exception:
            pass
        raise RuntimeError("Calendar overlay did not finish rendering (see calendar_dump_before_ready.html).")

    target_caption = f"{month_full} {year}"
    blocks = page.locator("div.rdp-month")
    target_mb = None
    for i in range(min(2, blocks.count())):
        mb = blocks.nth(i)
        cap = (mb.locator(".Calendar_month-text__eEKDO").inner_text() or "").strip()
        if cap == target_caption:
            target_mb = mb
            break

    next_btn = page.locator("[data-uitest='next-month']").first
    step = 0
    while not target_mb and step < 12:
        if (next_btn.get_attribute("aria-disabled") or "").lower() == "true":
            break
        try:
            next_btn.click(timeout=1200)
        except Exception:
            try: next_btn.click(force=True, timeout=1200)
            except Exception: break
        _wait_calendar_ready(page, timeout_ms=4000)
        blocks = page.locator("div.rdp-month")
        for i in range(min(2, blocks.count())):
            mb = blocks.nth(i)
            cap = (mb.locator(".Calendar_month-text__eEKDO").inner_text() or "").strip()
            if cap == target_caption:
                target_mb = mb
                break
        step += 1

    if not target_mb:
        target_mb = blocks.first

    if not _try_click_range(page, target_mb, int(day)):
        other = blocks.nth(1) if target_mb == blocks.nth(0) else blocks.nth(0)
        if other.count() and _try_click_range(page, other, int(day)):
            pass
        else:
            all_enabled = page.locator("div.rdp-month button[class*='rdp-day']:not([disabled])")
            if all_enabled.count() >= 2:
                try: all_enabled.nth(0).click(timeout=1000)
                except Exception: all_enabled.nth(0).click(force=True, timeout=1000)
                try: all_enabled.nth(1).click(timeout=1000)
                except Exception: all_enabled.nth(1).click(force=True, timeout=1000)
            else:
                page.screenshot(path="date_day_not_found.png")
                Path("calendar_dump.html").write_text(
                    page.locator("#range-picker, .rdp").first.evaluate("el=>el.outerHTML"),
                    encoding="utf-8"
                )
                raise RuntimeError("No enabled dates to select a range. See date_day_not_found.png / calendar_dump.html.")

    caps = [(page.locator(".Calendar_month-text__eEKDO").nth(i).inner_text() or "").strip()
            for i in range(page.locator(".Calendar_month-text__eEKDO").count())]
    print(f"📅 Selected a range (caps visible after selection: {caps})")

def _find_confirm_locators(page):
    cands = []
    cands.append(page.locator("button[data-uitest='select-check-in-cta']").first)
    cands.append(page.get_by_role("button", name=re.compile(r"^Confirm Dates$", re.I)).first)
    cands.append(page.locator("button:has-text('Confirm Dates')").first)
    cands.append(page.locator("text=Confirm Dates").first)
    return cands

def finalize_dates_in_picker(page, timeout_ms=12000):
    # 1) Done (inside the calendar)
    done = page.get_by_role("button", name=re.compile(r"^Done$", re.I)).first
    if done.count():
        start = time.time()
        while time.time() - start < 6:
            try:
                if done.is_enabled(): break
            except Exception:
                pass
            time.sleep(0.2)
        try:
            done.click(timeout=1500)
        except Exception:
            done.click(force=True, timeout=1500)

    # 2) Wait for outer Confirm to enable
    start = time.time()
    while (time.time() - start) * 1000 < timeout_ms:
        for cand in _find_confirm_locators(page):
            try:
                if cand.count() and cand.is_visible() and cand.is_enabled():
                    return True
            except Exception:
                pass
        time.sleep(0.2)

    try:
        Path("confirm_area_dump.html").write_text(page.content(), encoding="utf-8")
    except Exception:
        pass
    page.screenshot(path="confirm_disabled.png")
    raise RuntimeError("Confirm Dates (outer) did not enable. See confirm_disabled.png / confirm_area_dump.html.")

def click_confirm_dates(page, timeout_ms=12000):
    cands = _find_confirm_locators(page)
    cta = None
    end = time.time() + (timeout_ms / 1000.0)
    while time.time() < end and cta is None:
        for cand in cands:
            try:
                if cand.count() and cand.is_visible() and cand.is_enabled():
                    cta = cand
                    break
            except Exception:
                pass
        time.sleep(0.2)

    if not cta:
        try:
            Path("confirm_area_dump.html").write_text(page.content(), encoding="utf-8")
        except Exception:
            pass
    else:
        try:
            cta.click(timeout=1500)
        except Exception:
            cta.click(force=True, timeout=1500)

    # Ensure calendar overlay is gone before moving on
    overlay = page.locator("#range-picker, .rdp").first
    try:
        overlay.wait_for(state="hidden", timeout=6000)
    except Exception:
        pass

# ---------- Fill form ----------
def fill_booking_form(page, month, year, _checkin_full_aria_unused):
    set_counter(page, "adults", ADULTS)
    set_counter(page, "children", CHILDREN)
    set_unit_sizes(page, UNIT_SIZES)

    nights_input = page.locator("#number-of-nights")
    if nights_input.count():
        nights_input.click()
        nights_input.fill(str(NIGHTS))

    pick_month(page, month)
    pick_year(page, year)
    pick_checkin_date(page, month, year, CHECKIN_DAY)

# ---------- Results scraping ----------
MONTH_NAME_RE = r"(January|February|March|April|May|June|July|August|September|October|November|December)"

def extract_date_range(page) -> str:
    """
    Returns text like: 'January 2 - January 9, 2026'
    """
    try:
        bar = page.get_by_text(re.compile(r"Showing availability for", re.I)).first
        full = (bar.inner_text() or "").strip()
    except Exception:
        try:
            full = page.inner_text("body")
        except Exception:
            full = ""

    m = re.search(rf"{MONTH_NAME_RE} \d{{1,2}}\s*-\s*{MONTH_NAME_RE} \d{{1,2}}, \d{{4}}", full)
    return m.group(0).strip() if m else ""

def _normalize_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def clean_points_from_text(txt: str) -> str:
    """
    Grabs the first '<number> points' occurrence from a block of text.
    """
    if not txt:
        return ""
    m = re.search(r"([\d,]+)\s*points", txt, re.I)
    if m:
        return m.group(1).replace(",", "")
    # fallback: any number that looks like points (avoid $ and %)
    m2 = re.search(r"\b(\d{1,3}(?:,\d{3})+|\d{4,})\b", txt)
    return m2.group(1).replace(",", "") if m2 else ""

def scroll_results_into_view(page, max_scrolls=20, pause=0.4):
    """Progressively scrolls the results page to trigger lazy loading."""
    try:
        last_h = page.evaluate("() => document.body.scrollHeight")
    except Exception:
        last_h = None
    for _ in range(max_scrolls):
        page.mouse.wheel(0, 2500)
        time.sleep(pause)
        try:
            page.wait_for_load_state("networkidle", timeout=1500)
        except Exception:
            pass
        try:
            new_h = page.evaluate("() => document.body.scrollHeight")
        except Exception:
            new_h = last_h
        if new_h == last_h:
            break
        last_h = new_h

def wait_for_result_cards(page, timeout_ms=25000):
    """Wait specifically for cards/rows that contain 'points' to exist."""
    sel = (
        "[data-uitest='availability-result-card'], "
        "[data-uitest='availability-resort-result'], "
        "section:has(h2), article:has(h2)"
    )
    end = time.time() + (timeout_ms / 1000.0)
    while time.time() < end:
        try:
            cards = page.locator(sel)
            if cards.count() > 0:
                with_points = cards.filter(has_text=re.compile(r"\bpoints\b", re.I))
                if with_points.count() > 0:
                    return True
        except Exception:
            pass
        time.sleep(0.4)
    return False

def scrape_results(page):
    """
    Collect rows: Date Range, Resort, Room, Points
    Super-tolerant: tries multiple card shapes and falls back to block text parsing.
    """
    date_range = extract_date_range(page)
    rows = []

    # Ensure everything is rendered
    scroll_results_into_view(page, max_scrolls=24, pause=0.35)
    wait_for_result_cards(page, timeout_ms=12000)

    def _norm(s: str) -> str:
        return re.sub(r"\s+", " ", (s or "").strip())

    def _points_from_text(txt: str) -> str:
        if not txt:
            return ""
        m = re.search(r"([\d,]+)\s*points", txt, re.I)
        if m:
            return m.group(1).replace(",", "")
        m2 = re.search(r"\b(\d{1,3}(?:,\d{3})+|\d{4,})\b", txt)
        return m2.group(1).replace(",", "") if m2 else ""

    # Primary: explicit “result” cards
    card_sel = (
        "[data-uitest='availability-result-card'], "
        "[data-uitest='availability-resort-result']"
    )
    cards = page.locator(card_sel)
    if cards.count() > 0:
        for i in range(cards.count()):
            card = cards.nth(i)
            try:
                text = card.inner_text()
            except Exception:
                text = ""
            # Resort
            resort = ""
            for rsel in ("[data-uitest='resort-name']", "h2", "[class*='resort'] h2"):
                try:
                    loc = card.locator(rsel).first
                    if loc.count():
                        resort = _norm(loc.inner_text()); 
                        if resort: break
                except Exception:
                    pass
            # Room
            room = ""
            for usel in ("[data-uitest*='unit-name']", "h3", "[class*='unit'] h3", "h4"):
                try:
                    loc = card.locator(usel).first
                    if loc.count():
                        room = _norm(loc.inner_text()); 
                        if room: break
                except Exception:
                    pass
            pts = _points_from_text(text)
            if resort or room or pts:
                rows.append({"Date Range": date_range, "Resort": resort, "Room": room, "Points": pts})

    # Secondary: section/article blocks that *mention points*
    if not rows:
        blocks = page.locator("section:has(h2), article:has(h2)")
        blocks = blocks.filter(has_text=re.compile(r"\bpoints\b", re.I))
        for i in range(blocks.count()):
            b = blocks.nth(i)
            try:
                text = b.inner_text()
            except Exception:
                text = ""
            resort = ""
            for rsel in ("h2",):
                try:
                    loc = b.locator(rsel).first
                    if loc.count():
                        resort = _norm(loc.inner_text()); 
                        if resort: break
                except Exception:
                    pass
            room = ""
            for usel in ("[data-uitest*='unit-name']", "h3", "h4"):
                try:
                    loc = b.locator(usel).first
                    if loc.count():
                        room = _norm(loc.inner_text()); 
                        if room: break
                except Exception:
                    pass
            pts = _points_from_text(text)
            if resort or room or pts:
                rows.append({"Date Range": date_range, "Resort": resort, "Room": room, "Points": pts})

    # Tertiary: brute force — find any nearby heading for nodes containing "points"
    if not rows:
        pts_nodes = page.locator("xpath=//*[contains(translate(., 'POINTS','points'),'points')]")
        count = min(60, pts_nodes.count())
        for i in range(count):
            node = pts_nodes.nth(i)
            try:
                block = node.locator("xpath=ancestor::section[1] | xpath=ancestor::article[1] | xpath=ancestor::div[1]").first
            except Exception:
                continue
            if not block.count():
                continue
            try:
                text = block.inner_text()
            except Exception:
                text = ""
            pts = _points_from_text(text)
            if not pts:
                continue
            resort = ""
            for rsel in ("h2", "header h2"):
                loc = block.locator(rsel).first
                if loc.count():
                    resort = _norm(loc.inner_text()); 
                    if resort: break
            room = ""
            for usel in ("[data-uitest*='unit-name']", "h3", "h4"):
                loc = block.locator(usel).first
                if loc.count():
                    room = _norm(loc.inner_text()); 
                    if room: break
            rows.append({"Date Range": date_range, "Resort": resort, "Room": room, "Points": pts})

    # Dedup + sanity
    seen = set()
    out = []
    for r in rows:
        key = (r["Date Range"], r["Resort"], r["Room"], r["Points"])
        if key in seen:
            continue
        seen.add(key)
        out.append(r)

    print(f"🧾 Parsed rows: {len(out)}")
    if len(out) == 0:
        # Save for inspection if still empty
        try:
            Path("no_results_debug.html").write_text(page.content(), encoding="utf-8")
        except Exception:
            pass
        page.screenshot(path="no_results_debug.png", full_page=True)
        print("⚠️ No rows parsed. Dumped no_results_debug.(html|png) for review.")
    return out

# ---------- Saving ----------
def save_csv(rows, csv_path: Path):
    new_file = not csv_path.exists()
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Date Range", "Resort", "Room", "Points"])
        if new_file:
            writer.writeheader()
        for r in rows:
            writer.writerow(r)

def save_xlsx(rows, xlsx_path: Path):
    try:
        from openpyxl import Workbook, load_workbook
    except Exception:
        print("ℹ️ openpyxl not installed; skipping Excel output.")
        return

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Date Range", "Resort", "Room", "Points"])

    for r in rows:
        ws.append([r["Date Range"], r["Resort"], r["Room"], r["Points"]])

    wb.save(xlsx_path)

# ----------------- Main -----------------
def scrape_availability_member():
    month, year, checkin_full_aria = prompt_month_year()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS, slow_mo=SLOW_MO)
        context = browser.new_context(viewport={"width": 1400, "height": 900})
        page = context.new_page()
        page.set_default_timeout(15000)

        # Login
        page.goto("https://holidayinnclub.com/login", wait_until="domcontentloaded")
        page.fill("#okta-signin-username", USERNAME)
        page.fill("#okta-signin-password", PASSWORD)
        page.click("#okta-signin-submit")

        # Wait for either dashboard or "Book My Vacation"
        try:
            page.get_by_role("button", name=re.compile(r"Book My Vacation", re.I)).wait_for(timeout=20000)
        except Exception:
            page.wait_for_load_state("networkidle", timeout=20000)

        print("✅ Logged in.")

        # Book My Vacation (may open same tab or new tab)
        new_tab = None
        try:
            with context.expect_page(timeout=5000) as newp:
                try:
                    page.get_by_role("button", name=re.compile(r"^Book My Vacation$", re.I)).click()
                except Exception:
                    page.get_by_role("link", name=re.compile(r"^Book My Vacation$", re.I)).click()
            new_tab = newp.value
        except Exception:
            pass
        if new_tab:
            page = new_tab
        page.wait_for_load_state("domcontentloaded")

        wait_for_booking_form(page)
        accept_cookies_if_present(page)

        # Florida selection
        page.get_by_text(re.compile(r"Location\b", re.I)).scroll_into_view_if_needed()
        select_all_florida_locations(page)

        # Fill rest + date range
        fill_booking_form(page, month, year, checkin_full_aria)

        # Finish the date flow: Done -> Confirm
        finalize_dates_in_picker(page, timeout_ms=12000)
        click_confirm_dates(page, timeout_ms=12000)

        # Let results settle, then force-load all cards
        try:
            page.wait_for_load_state("networkidle", timeout=8000)
        except Exception:
            pass
        scroll_results_into_view(page, max_scrolls=24, pause=0.35)

        # Results
        wait_for_results(page)
        write_debug(page, "member_results")
        print("📸 Saved screenshot + HTML of member results.")

        rows = scrape_results(page)
        if rows:
            save_csv(rows, CSV_PATH)
            save_xlsx(rows, XLSX_PATH)
            print(f"✅ Appended {len(rows)} rows to {CSV_PATH}")
            if XLSX_PATH.exists():
                print(f"✅ Also wrote Excel: {XLSX_PATH}")
        else:
            print("⚠️ No results parsed on the results page.")
            write_debug(page, "no_results_debug")

        time.sleep(1)
        browser.close()

if __name__ == "__main__":
    scrape_availability_member()
